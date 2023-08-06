import requests
from bs4 import BeautifulSoup
import openpyxl

fpath = r'/Users/lily/crawling/Excel/네이버주식정보_data.xlsx'

# 1) 엑셀 만들기
wb = openpyxl.Workbook()

# 2) 엑셀 워크시트 만들기
ws = wb.create_sheet('네이버 주식 정보')

# 3) 데이터 추가하기
ws['A1'] = '종목'
ws['B1'] = '현재가'
ws['C1'] = '고가'
ws['D1'] = '저가'
ws['E1'] = '거래량'

# 종목 코드 리스트
codes = [
    '005930',
    '000660',
    '035720'
]

for index, code in enumerate(codes):
    i = index + 2
    url = f"https://finance.naver.com/item/sise.naver?code={code}"
    response = requests.get(url)
    html = response.text
    soup = BeautifulSoup(html, 'html.parser')
    # 종목 명
    name = soup.select_one('div.wrap_company>h2>a').text
    ws[f'A{i}'] = name

    # 현재가
    nowPrice = soup.select_one("#_nowVal").text
    nowPrice = nowPrice.replace(',', '')
    ws[f'B{i}'] = nowPrice

    # 고가
    highPrice = soup.select_one("#_high").text
    highPrice = highPrice.replace(',', '')
    ws[f'C{i}'] = highPrice

    # 저가
    lowPrice = soup.select_one("#_low").text
    lowPrice = lowPrice.replace(',', '')
    ws[f'D{i}'] = lowPrice

    # 거래량
    tradingVolume = soup.select_one("#_quant").text
    tradingVolume = tradingVolume.replace(',', '')
    ws[f'E{i}'] = tradingVolume

wb.save(fpath)